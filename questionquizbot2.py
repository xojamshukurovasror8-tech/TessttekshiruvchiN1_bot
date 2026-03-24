import os
import re
import json
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters,
)
import openpyxl

# =======================================
# JSONNI HAR SAFAR O‘QISH
# =======================================
def load_tests():
    try:
        with open("asrorquestions1.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print("JSON o‘qishda xatolik:", e)
        return {}

# =======================================
# GURUHLAR VA VAQT
# =======================================
GRUPPLAR = {
    "8:30 TOQ": {"raqam": "toq", "vaqt": "8:30"},
    "10:00 TOQ": {"raqam": "juft", "vaqt": "8:30"},
    "16:00 TOQ": {"raqam": "toq", "vaqt": "9:30"},
    "17:30 TOQ": {"raqam": "juft", "vaqt": "9:30"},
    "8:30 JUFT": {"raqam": "toq", "vaqt": "10:30"},
    "17:30 JUFT": {"raqam": "juft", "vaqt": "10:30"},
    "19:00 JUFT": {"raqam": "toq", "vaqt": "11:30"},
}

JAVOB, ISM, GURUH = range(3)

# =======================================
# EXCELGA SAQLASH
# =======================================
def save_to_excel_grouped(ism, guruh, togri, jami, foiz):
    filename = "results_grouped.xlsx"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Natijalar"
        ws.append(["Guruh", "Ism", "To‘g‘ri", "Ball", "Foiz", "Vaqt"])
        wb.save(filename)
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    guruh_rows = [cell.value for cell in ws["A"] if cell.value == guruh]
    if not guruh_rows:
        ws.append([])

    ws.append([guruh, ism, togri, togri, foiz, now])
    wb.save(filename)

# =======================================
# BOT FUNKSIYALARI
# =======================================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [[InlineKeyboardButton("Test javoblarini tekshirish", callback_data="start_test")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        "Assalomu alaykum! 👋\nQuyidagi tugma orqali test javoblarini yuborishingiz mumkin.",
        reply_markup=reply_markup
    )
    return ConversationHandler.END

async def start_test_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.reply_text("👤 Ism va Familiyangizni to'liq  kiriting:")
    return ISM

async def get_ism(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["ism"] = update.message.text.strip()
    await update.message.reply_text("🏫 Guruhingizni kiriting (masalan 8:30 TOQ):")
    return GURUH

async def get_guruh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    guruh = update.message.text.strip().upper()
    if guruh not in GRUPPLAR:
        await update.message.reply_text("❌ Bunday guruh topilmadi. Qaytadan kiriting:")
        return GURUH
    context.user_data["guruh"] = guruh
    await update.message.reply_text(
        "📘 Javoblarni yuboring formatda:\n"
        "[   BERILGAN TEST KODI   ]+[   JAVOBLAR    ]\n"
        "Misol: test-1+1B2D3C..."
    )
    return JAVOB

async def tekshir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    if "+" not in text:
        await update.message.reply_text("❗ Format noto‘g‘ri. Masalan: test-1+1A2B3C…")
        return JAVOB

    test_nomi, user_javob_raw = text.split("+", 1)

    # 🔥 HAR SAFAR JSON QAYTA O‘QILADI
    TESTLAR = load_tests()

    if test_nomi not in TESTLAR:
        await update.message.reply_text("❗ Bunday test topilmadi")
        return JAVOB

    javoblar_list = re.findall(r"(\d+)([a-dA-D])", user_javob_raw)
    savollar = TESTLAR[test_nomi]["savollar"]

    togri = 0
    xatolar = []

    for num, variant in javoblar_list:
        num_str = str(num)
        variant = variant.upper()

        if num_str in savollar:
            if variant == savollar[num_str].upper():
                togri += 1
            else:
                xatolar.append(f"{num}-{savollar[num_str]}")
        else:
            xatolar.append(f"{num}-??")

    jami = len(savollar)
    foiz = int((togri / jami) * 100)

    await update.message.reply_text(
        f"👤 Ism: {context.user_data.get('ism')}\n"
        f"🏫 Guruh: {context.user_data.get('guruh')}\n"
        f"📘 Test: {test_nomi}\n"
        f"✅ To‘g‘ri: {togri}/{jami}\n"
        f"📊 Foiz: {foiz}%\n"
        + ("\n❌ To‘g‘ri javoblar: " + ", ".join(xatolar) if xatolar else "\n🎉 Barchasi to‘g‘ri!")
    )

    save_to_excel_grouped(
        ism=context.user_data.get("ism"),
        guruh=context.user_data.get("guruh"),
        togri=togri,
        jami=jami,
        foiz=foiz
    )

    return ConversationHandler.END

async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    filename = "results_grouped.xlsx"
    if not os.path.exists(filename):
        await update.message.reply_text("Hech kim javob yozmagan")
        return

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    noyob = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        ism, guruh = row[1], row[0]
        noyob.add((ism, guruh))

    await update.message.reply_text(f"👥 Javob bergan foydalanuvchilar soni: {len(noyob)}")

# =======================================
# BOT OCHISH
# =======================================
app = ApplicationBuilder().token("8508804567:AAF4VLi1wW1uYcs_rJV55CwD6BcEaj8Xj3A").build()

conv_handler = ConversationHandler(
    entry_points=[CallbackQueryHandler(start_test_callback, pattern="start_test")],
    states={
        ISM: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_ism)],
        GURUH: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_guruh)],
        JAVOB: [MessageHandler(filters.TEXT & ~filters.COMMAND, tekshir)],
    },
    fallbacks=[],
)

app.add_handler(CommandHandler("start", start))
app.add_handler(conv_handler)
app.add_handler(CommandHandler("stats", stats))

app.run_polling()
